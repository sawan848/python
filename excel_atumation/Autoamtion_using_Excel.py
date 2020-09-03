import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def process_workbook(filename):
    workbook=xl.load_workbook(filename)
    sheet=workbook['Sheet1']
    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        corrected_price=cell.value*0.9
        #creating a new cell
        corrected_price_cell=sheet.cell(row,4)
        corrected_price_cell.value=corrected_price

        values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
#values object have the all the values in the column D(4)

# create a instance of a barchart(load a barchart)
        chart=BarChart()
        chart.add_data(values)
        sheet.add_chart(chart,'e2')

        workbook.save(filename)

print('Hello')
#for further detail see the exceprion_handling.ipynb
"""
this file is the example automation (create,modified )files using a xml
1.firstly we !pip install openpyxl
2.load the xml file
3.import the barchar and reference
4.save a new file name😊
"""